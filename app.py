from flask import Flask, request, render_template, send_file, flash, redirect, url_for
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import os
import tempfile
from pathlib import Path
import zipfile
from werkzeug.utils import secure_filename
import io

app = Flask(__name__)
app.secret_key = 'consolidador_excel_2024'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configurações
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

# Criar pasta de uploads se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def criar_estilos():
    """
    Cria e retorna os estilos personalizados para o relatório
    """
    # Borda cinza claro para a tabela de classificação
    borda_cinza = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0')
    )
    
    # Estilo para o cabeçalho principal
    estilo_cabecalho_principal = NamedStyle(name='estilo_cabecalho_principal')
    estilo_cabecalho_principal.font = Font(name='Calibri', size=15, bold=True)
    estilo_cabecalho_principal.alignment = Alignment(horizontal='center', vertical='center')
    estilo_cabecalho_principal.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    estilo_cabecalho_principal.border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Estilo para labels principais (fonte maior)
    estilo_label_principal = NamedStyle(name='estilo_label_principal')
    estilo_label_principal.font = Font(name='Calibri', size=12)
    estilo_label_principal.alignment = Alignment(horizontal='left', vertical='center')
    
    # Estilo para labels (sem bordas)
    estilo_label = NamedStyle(name='estilo_label')
    estilo_label.font = Font(name='Calibri', size=11)
    estilo_label.alignment = Alignment(horizontal='left', vertical='center')
    
    # Estilo para valores (sem bordas) - alinhado à esquerda
    estilo_valor = NamedStyle(name='estilo_valor')
    estilo_valor.font = Font(name='Calibri', size=11)
    estilo_valor.alignment = Alignment(horizontal='left', vertical='center')
    
    # Estilo para percentuais (sem bordas) - alinhado à esquerda
    estilo_percentual = NamedStyle(name='estilo_percentual')
    estilo_percentual.font = Font(name='Calibri', size=11)
    estilo_percentual.alignment = Alignment(horizontal='left', vertical='center')
    estilo_percentual.number_format = '0.00000%'
    
    # Estilo para cabeçalho da tabela de classificação (negrito, centralizado, com bordas)
    estilo_cabecalho_tabela = NamedStyle(name='estilo_cabecalho_tabela')
    estilo_cabecalho_tabela.font = Font(name='Calibri', size=11, bold=True)
    estilo_cabecalho_tabela.alignment = Alignment(horizontal='center', vertical='center')
    estilo_cabecalho_tabela.border = borda_cinza
    
    # Estilo para linhas da tabela de classificação (centralizado, com bordas)
    estilo_linha_tabela = NamedStyle(name='estilo_linha_tabela')
    estilo_linha_tabela.font = Font(name='Calibri', size=11)
    estilo_linha_tabela.alignment = Alignment(horizontal='center', vertical='center')
    estilo_linha_tabela.border = borda_cinza
    
    return {
        'cabecalho_principal': estilo_cabecalho_principal,
        'label_principal': estilo_label_principal,
        'label': estilo_label,
        'valor': estilo_valor,
        'percentual': estilo_percentual,
        'cabecalho_tabela': estilo_cabecalho_tabela,
        'linha_tabela': estilo_linha_tabela
    }

def converter_numero_br(valor):
    """
    Converte um número em formato brasileiro (vírgula) para float
    """
    if isinstance(valor, str):
        # Remove % se existir
        valor = valor.strip('%')
        # Substitui vírgula por ponto
        valor = valor.replace(',', '.')
        try:
            return float(valor)
        except ValueError:
            return valor
    return valor

def processar_dados_planilha(df):
    """
    Processa o DataFrame para extrair os dados necessários no formato correto
    """
    dados = {}
    
    # Função auxiliar para encontrar valor em uma coluna específica
    def encontrar_valor(df, texto_busca):
        for idx, row in df.iterrows():
            for col in df.columns:
                valor = row[col]
                if pd.notna(valor) and isinstance(valor, str) and texto_busca.lower() in valor.lower():
                    # Procurar o valor na próxima coluna ou célula
                    for next_col in df.columns[df.columns.get_loc(col):]:
                        next_valor = row[next_col]
                        if pd.notna(next_valor) and next_valor != valor:
                            return next_valor
        return None
    
    # Extrair dados específicos
    dados['escola'] = encontrar_valor(df, 'Identificação da Escola')
    dados['classificacao'] = encontrar_valor(df, 'Classificação Geral')
    dados['pontuacao'] = encontrar_valor(df, 'Pontuação Geral')
    
    # Extrair pontuações dos blocos
    blocos = {
        'bloco1': '1. Edifícios e Instalações da Área de Preparo de Alimentos:',
        'bloco2': '2. Equipamentos para Temperatura Controlada:',
        'bloco3': '3. Manipuladores:',
        'bloco4': '4. Recebimento:',
        'bloco5': '5. Processos e Produções:',
        'bloco6': '6. Higienização Ambiental:'
    }
    
    for bloco_key, texto_busca in blocos.items():
        dados[bloco_key] = encontrar_valor(df, texto_busca)
    
    return dados

def criar_layout_planilha(ws, dados, estilos):
    """
    Cria o layout exato da planilha com os dados fornecidos
    """
    # Remover linhas de grade
    ws.sheet_view.showGridLines = False
    
    # Cabeçalho principal na primeira linha
    ws['A1'] = 'Relatório - Lista de Verificação em Boas Práticas'
    ws['A1'].style = estilos['cabecalho_principal']
    ws.merge_cells('A1:P1')
    
    # Estrutura fixa - Labels principais (com fonte maior)
    estrutura_principal = [
        ('D4', 'Identificação da Escola:', 'label_principal'),
        ('D6', 'Classificação Geral:', 'label_principal'),
        ('D8', 'Pontuação Geral:', 'label_principal'),
        ('D10', 'Classificação por Bloco:', 'label_principal')
    ]
    
    # Aplicar estrutura principal
    for celula, texto, estilo in estrutura_principal:
        ws[celula] = texto
        ws[celula].style = estilos[estilo]
    
    # Estrutura dos blocos com células mescladas
    blocos = [
        ('D12', 'H12', '1. Edifícios e Instalações da Área de Preparo de Alimentos:', 'L12'),
        ('D14', 'H14', '2. Equipamentos para Temperatura Controlada:', 'L14'),
        ('D16', 'H16', '3. Manipuladores:', 'L16'),
        ('D18', 'H18', '4. Recebimento:', 'L18'),
        ('D20', 'H20', '5. Processos e Produções:', 'L20'),
        ('D22', 'H22', '6. Higienização Ambiental:', 'L22')
    ]
    
    # Aplicar blocos e mesclar células
    for inicio, fim, texto, celula_valor in blocos:
        ws[inicio] = texto
        ws[inicio].style = estilos['label']
        # Mesclar células do texto
        ws.merge_cells(f'{inicio}:{fim}')
    
    # Mapear dados nas células corretas com mesclagem
    mapeamento_dados = [
        ('G4', 'O4', dados['escola'], 'valor'),        # G4:O4 mesclado
        ('G6', 'O6', dados['classificacao'], 'valor'), # G6:O6 mesclado
        ('G8', 'O8', dados['pontuacao'], 'percentual'), # G8:O8 mesclado
        ('L12', None, dados['bloco1'], 'percentual'),
        ('L14', None, dados['bloco2'], 'percentual'),
        ('L16', None, dados['bloco3'], 'percentual'),
        ('L18', None, dados['bloco4'], 'percentual'),
        ('L20', None, dados['bloco5'], 'percentual'),
        ('L22', None, dados['bloco6'], 'percentual')
    ]
    
    # Aplicar dados
    for celula_inicio, celula_fim, valor, estilo in mapeamento_dados:
        if valor is not None:
            if estilo == 'percentual':
                # Converter para decimal
                valor_numerico = converter_numero_br(valor)
                if isinstance(valor_numerico, float):
                    ws[celula_inicio].value = valor_numerico / 100  # Converter para decimal
                    ws[celula_inicio].number_format = '0.00000%'
                else:
                    ws[celula_inicio].value = valor
            else:
                ws[celula_inicio].value = valor
            
            ws[celula_inicio].style = estilos[estilo]
            
            # Mesclar células se necessário
            if celula_fim:
                ws.merge_cells(f'{celula_inicio}:{celula_fim}')
    
    # Tabela de classificação com células mescladas (cabeçalhos)
    classificacao_cabecalho = [
        ('D24', 'Classificação', 'cabecalho_tabela'),
    ]
    
    # Tabela de classificação com células mescladas (linhas)
    classificacao_linhas = [
        ('D25', 'Situação de risco sanitário muito alto', 'linha_tabela'),
        ('D26', 'Situação de risco sanitário alto', 'linha_tabela'),
        ('D27', 'Situação de risco sanitário regular', 'linha_tabela'),
        ('D28', 'Situação de risco sanitário baixo', 'linha_tabela'),
        ('D29', 'Situação de risco sanitário muito baixo', 'linha_tabela')
    ]
    
    # Aplicar cabeçalho da classificação
    for celula, texto, estilo in classificacao_cabecalho:
        ws[celula] = texto
        ws[celula].style = estilos[estilo]
        # Mesclar células (D até G)
        inicio_merge = celula  # Ex: D24
        fim_merge = 'G' + celula[1:]  # Ex: G24
        ws.merge_cells(f'{inicio_merge}:{fim_merge}')
    
    # Aplicar linhas da classificação
    for celula, texto, estilo in classificacao_linhas:
        ws[celula] = texto
        ws[celula].style = estilos[estilo]
        # Mesclar células (D até G)
        inicio_merge = celula  # Ex: D25
        fim_merge = 'G' + celula[1:]  # Ex: G25
        ws.merge_cells(f'{inicio_merge}:{fim_merge}')
    
    # Tabela de pontuação com células mescladas (cabeçalho)
    pontuacao_cabecalho = [
        ('H24', 'Pontuação (%)', 'cabecalho_tabela'),
    ]
    
    # Tabela de pontuação com células mescladas (linhas)
    pontuacao_linhas = [
        ('H25', '0 a 25', 'linha_tabela'),
        ('H26', '26 a 50', 'linha_tabela'),
        ('H27', '51 a 75', 'linha_tabela'),
        ('H28', '76 a 90', 'linha_tabela'),
        ('H29', '90 a 100', 'linha_tabela')
    ]
    
    # Aplicar cabeçalho da pontuação
    for celula, texto, estilo in pontuacao_cabecalho:
        ws[celula] = texto
        ws[celula].style = estilos[estilo]
        # Mesclar células (H até L)
        inicio_merge = celula  # Ex: H24
        fim_merge = 'L' + celula[1:]  # Ex: L24
        ws.merge_cells(f'{inicio_merge}:{fim_merge}')
    
    # Aplicar linhas da pontuação
    for celula, texto, estilo in pontuacao_linhas:
        ws[celula] = texto
        ws[celula].style = estilos[estilo]
        # Mesclar células (H até L)
        inicio_merge = celula  # Ex: H25
        fim_merge = 'L' + celula[1:]  # Ex: L25
        ws.merge_cells(f'{inicio_merge}:{fim_merge}')
    
    # Ocultar colunas específicas
    colunas_ocultas = ['B', 'C', 'E', 'I', 'K']
    for col in colunas_ocultas:
        ws.column_dimensions[col].hidden = True
    
    # Ajustar largura das colunas
    # Coluna A com 10 pontos
    ws.column_dimensions['A'].width = 10
    
    # Coluna D com 25 pontos
    ws.column_dimensions['D'].width = 25
    
    # Outras colunas visíveis
    outras_colunas = {'F': 15, 'G': 15, 'H': 5, 'J': 3, 'L': 15, 'M': 15, 'N': 15, 'O': 15, 'P': 15}
    for col, width in outras_colunas.items():
        ws.column_dimensions[col].width = width
    
    # Ajustar altura das linhas
    for row in range(1, 30):
        ws.row_dimensions[row].height = 20
    
    # Altura específica para a primeira linha (cabeçalho)
    ws.row_dimensions[1].height = 50
    
    # Definir altura específica para as linhas 13, 15, 17, 19, 21 (0,5)
    linhas_pequenas = [2, 5, 7, 13, 15, 17, 19, 21]
    for linha in linhas_pequenas:
        ws.row_dimensions[linha].height = 0.5

def consolidar_arquivos_web(arquivos_paths):
    """
    Consolida arquivos Excel mantendo formatação específica
    """
    try:
        # Criar novo workbook para consolidação
        wb_consolidado = Workbook()
        # Remover a planilha padrão
        wb_consolidado.remove(wb_consolidado.active)
        
        # Criar estilos uma vez para todo o workbook
        estilos = criar_estilos()
        
        # Registrar estilos no workbook
        for estilo in estilos.values():
            if estilo.name not in wb_consolidado._named_styles:
                wb_consolidado.add_named_style(estilo)
        
        for arquivo_path in arquivos_paths:
            try:
                # Obter nome do arquivo sem extensão
                nome_arquivo = Path(arquivo_path).stem
                
                # Ler o arquivo Excel
                df_dict = pd.read_excel(
                    arquivo_path,
                    sheet_name=None,
                    engine='openpyxl' if arquivo_path.endswith('.xlsx') else None
                )
                
                # Para cada planilha no arquivo
                for nome_planilha_original, df in df_dict.items():
                    # Nome da planilha no arquivo consolidado
                    nome_planilha = nome_arquivo[:31]
                    
                    # Verificar se já existe planilha com esse nome
                    contador = 1
                    nome_original = nome_planilha
                    while nome_planilha in wb_consolidado.sheetnames:
                        nome_planilha = f"{nome_original}_{contador}"
                        contador += 1
                    
                    # Criar nova planilha
                    ws = wb_consolidado.create_sheet(nome_planilha)
                    
                    # Processar o DataFrame para extrair os dados necessários
                    dados_processados = processar_dados_planilha(df)
                    
                    # Aplicar formatação e dados
                    criar_layout_planilha(ws, dados_processados, estilos)
                    
            except Exception as e:
                print(f"Erro ao processar {arquivo_path}: {e}")
                continue
        
        return wb_consolidado
    
    except Exception as e:
        print(f"Erro geral na consolidação: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        flash('Nenhum arquivo selecionado', 'error')
        return redirect(url_for('index'))
    
    files = request.files.getlist('files[]')
    
    if not files or files[0].filename == '':
        flash('Nenhum arquivo selecionado', 'error')
        return redirect(url_for('index'))
    
    # Validar arquivos
    valid_files = []
    for file in files:
        if file and allowed_file(file.filename):
            valid_files.append(file)
        else:
            flash(f'Arquivo {file.filename} não é um formato válido (.xls, .xlsx)', 'error')
    
    if not valid_files:
        flash('Nenhum arquivo válido encontrado', 'error')
        return redirect(url_for('index'))
    
    try:
        # Salvar arquivos temporariamente
        temp_files = []
        for file in valid_files:
            filename = secure_filename(file.filename)
            temp_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(temp_path)
            temp_files.append(temp_path)
        
        # Consolidar arquivos
        wb_consolidado = consolidar_arquivos_web(temp_files)
        
        if wb_consolidado:
            # Salvar arquivo consolidado
            output_path = os.path.join(UPLOAD_FOLDER, 'Relatorio.xlsx')
            wb_consolidado.save(output_path)
            wb_consolidado.close()
            
            # Limpar arquivos temporários
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                except:
                    pass
            
            # Preparar informações para a página de resultado
            planilhas = wb_consolidado.sheetnames if hasattr(wb_consolidado, 'sheetnames') else []
            
            return render_template('resultado.html', 
                                 sucesso=True, 
                                 num_arquivos=len(valid_files),
                                 planilhas=planilhas)
        else:
            flash('Erro ao consolidar arquivos', 'error')
            return redirect(url_for('index'))
            
    except Exception as e:
        flash(f'Erro durante o processamento: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/download')
def download_file():
    output_path = os.path.join(UPLOAD_FOLDER, 'Relatorio.xlsx')
    
    if os.path.exists(output_path):
        return send_file(output_path, 
                        as_attachment=True, 
                        download_name='Relatorio_Consolidado.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        flash('Arquivo não encontrado', 'error')
        return redirect(url_for('index'))

@app.route('/health')
def health():
    return {'status': 'ok', 'message': 'Consolidador Excel funcionando'}

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 